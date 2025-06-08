from datetime import date, datetime, timedelta  
from flask import Blueprint, render_template, redirect, url_for, flash, send_file, jsonify, request, current_app
from markupsafe import Markup
from flask_login import login_required, current_user, login_user, logout_user
from app.models import User, Report, db, Product, MaterialCategory, MaterialNorm, MaterialConsumption, UserDashboard, MaterialReceipt, ProductionLog
from app.forms import UserForm, ReportForm, LoginForm, ComparePeriodsForm, MaterialCategoryForm, ProductForm, MaterialNormForm, ProduceProductForm, MaterialReceiptForm
from sqlalchemy import func
import matplotlib.pyplot as plt
import io
import os
import base64

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter 
from openpyxl.cell import MergedCell
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics 
from reportlab.lib.fonts import addMapping
from reportlab.pdfbase.ttfonts import TTFont 
from io import BytesIO


bp = Blueprint('routes', __name__)

@bp.route('/')
def home():
    return redirect(url_for('routes.dashboard'))

@bp.route('/dashboard')
@login_required
def dashboard():
    user_count = User.query.count()
 
    report_count = Report.query.count()
    
    return render_template('dashboard.html', 
                         user_count=user_count,
                         
                         report_count=report_count)

@bp.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        
        if user and user.check_password(form.password.data):
            login_user(user)
            flash('Вы успешно вошли в систему', 'success')
            next_page = request.args.get('next')
            return redirect(next_page or url_for('routes.dashboard'))
        else:
            flash('Неверное имя пользователя или пароль', 'danger')
    return render_template('login.html', form=form)

@bp.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Вы успешно вышли из системы', 'success')
    return redirect(url_for('routes.home'))

@bp.route('/admin/users')
@login_required
def admin_users():
    if current_user.role != 'admin':
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    users = User.query.all()
    return render_template('admin_users.html', users=users)

@bp.route('/admin/users/create', methods=['GET', 'POST'])
@login_required
def create_user():
    if current_user.role != 'admin':
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    form = UserForm()
    
    if form.validate_on_submit():
        try:
            user = User(
                username=form.username.data,
                role=form.role.data,
                full_name=form.full_name.data,
                contact_info=form.contact_info.data
            )
            user.set_password(form.password.data)
            db.session.add(user)
            db.session.commit()
            flash('Пользователь успешно создан', 'success')
            return redirect(url_for('routes.admin_users'))
        except Exception as e:
            db.session.rollback()
            flash(f'Ошибка при создании пользователя: {str(e)}', 'danger')
    
    return render_template('create_user.html', form=form)

@bp.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    if current_user.role != 'admin':
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    user = User.query.get_or_404(user_id)
    form = UserForm(obj=user)
    
    if form.validate_on_submit():
        user.username = form.username.data
        user.role = form.role.data
        user.full_name = form.full_name.data
        user.contact_info = form.contact_info.data
        
        if form.password.data:
            user.set_password(form.password.data)
        
        db.session.commit()
        flash('Пользователь успешно обновлен', 'success')
        return redirect(url_for('routes.admin_users'))
    
    return render_template('edit_user.html', form=form, user=user)

@bp.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@login_required
def delete_user(user_id):
    if current_user.role != 'admin':
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    user = User.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    flash('Пользователь успешно удален', 'success')
    return redirect(url_for('routes.admin_users'))

@bp.route('/production_journal')
@login_required
def production_journal():
    
    production_logs = ProductionLog.query.options(
        db.joinedload(ProductionLog.product), 
        db.joinedload(ProductionLog.user)
    ).order_by(ProductionLog.production_date.desc(), ProductionLog.created_at.desc()).all()
    
    total_production_log_cost = sum(pl.total_cost_calculated for pl in production_logs)
    
    return render_template('production_journal.html', 
                        production_logs=production_logs, 
                        total_production_log_cost=total_production_log_cost)

@bp.route('/reports/<int:report_id>/delete', methods=['POST'])
@login_required
def delete_report(report_id):
    if current_user.role not in ['admin', 'accountant']:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.report_list'))
    
    report = Report.query.get_or_404(report_id)
    
    db.session.delete(report)
    db.session.commit()
    
    flash('Отчет успешно удален', 'success')
    return redirect(url_for('routes.report_list'))

@bp.route('/reports')
@login_required
def report_list():
    if current_user.role not in ['admin', 'accountant', 'analyst']:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    reports = Report.query.order_by(Report.created_at.desc()).all()
    
    reports_data = []
    for report_item in reports: 
        data_item = {'report': report_item}
        
        if report_item.report_type == 'material_stock':
            
            data_item['item_count'] = len(report_item.config.get('materials', [])) if report_item.config else 0
            data_item['total_amount'] = None 
        reports_data.append(data_item)
    
    return render_template('report_list.html', reports_data=reports_data)

@bp.route('/reports/create_material_stock', methods=['POST'])
@login_required
def create_material_stock_report():
    if current_user.role not in ['admin', 'accountant', 'analyst']:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.report_list'))

    try:
        materials_snapshot = []
        all_materials = MaterialCategory.query.order_by(MaterialCategory.name).all()
        total_stock_value = 0
        for mat in all_materials:
            material_value = mat.stock_quantity * (mat.average_cost_price or 0)
            total_stock_value += material_value
            materials_snapshot.append({
                'id': mat.id,
                'code': mat.code,
                'name': mat.name,
                'unit': mat.unit,
                'stock_quantity': mat.stock_quantity,
                'average_cost_price': mat.average_cost_price or 0, 
                'total_value': material_value 
            })

        report_title = f"Отчет об остатках материалов на {datetime.now().strftime('%d.%m.%Y %H:%M')}"

        report = Report(
            report_type='material_stock',
            created_at=datetime.now(),
            user_id=current_user.id,
            config={ 
                'title': report_title,
                'materials': materials_snapshot,
                'total_stock_value': total_stock_value 
            }
        )
        db.session.add(report)
        db.session.commit()
        flash('Отчет об остатках материалов успешно сформирован!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при формировании отчета об остатках: {str(e)}', 'danger')
        current_app.logger.error(f"Material stock report generation error: {e}", exc_info=True)
    
    return redirect(url_for('routes.report_list'))

@bp.route('/reports/<int:report_id>/details')
@login_required
def report_details(report_id):
    if current_user.role not in ['admin', 'accountant', 'analyst']:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    report = Report.query.get_or_404(report_id)
    
    if report.report_type == 'material_stock':
        
        return render_template('report_details_stock.html', report=report, report_data=report.config)
    
    else:
        flash('Неизвестный или неподдерживаемый тип отчета.', 'warning')
        return redirect(url_for('routes.report_list'))
    
@bp.route('/reports/<int:report_id>/export/pdf')
@login_required
def export_report_pdf(report_id):
    if current_user.role not in ['admin', 'accountant', 'analyst']:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.report_list'))

    PDF_FONT_REGULAR_NAME = 'MyArial'
    PDF_FONT_BOLD_NAME = 'MyArial-Bold'
    PDF_FONT_ITALIC_NAME = 'MyArial-Italic'
    PDF_FONT_BOLDITALIC_NAME = 'MyArial-BoldItalic'

    font_to_use_regular = 'Helvetica'
    font_to_use_bold = 'Helvetica-Bold'

    path_regular_ttf = os.path.join(current_app.root_path, 'static', 'fonts', 'Arial.ttf') 
    path_bold_ttf = os.path.join(current_app.root_path, 'static', 'fonts', 'arialbd.ttf') 

    try:
        fonts_registered_successfully = False
        if os.path.exists(path_regular_ttf):
            pdfmetrics.registerFont(TTFont(PDF_FONT_REGULAR_NAME, path_regular_ttf))
            font_to_use_regular = PDF_FONT_REGULAR_NAME
            current_app.logger.info(f"Успешно зарегистрирован: {PDF_FONT_REGULAR_NAME} из {path_regular_ttf}")
            fonts_registered_successfully = True 

            if os.path.exists(path_bold_ttf):
                pdfmetrics.registerFont(TTFont(PDF_FONT_BOLD_NAME, path_bold_ttf))
                font_to_use_bold = PDF_FONT_BOLD_NAME
                current_app.logger.info(f"Успешно зарегистрирован: {PDF_FONT_BOLD_NAME} из {path_bold_ttf}")
            else:
                current_app.logger.warning(f"НЕ НАЙДЕН файл жирного шрифта: {path_bold_ttf}. Для жирного будет использован: {font_to_use_bold}")

            if font_to_use_regular == PDF_FONT_REGULAR_NAME and font_to_use_bold == PDF_FONT_BOLD_NAME:

                addMapping(PDF_FONT_REGULAR_NAME, 
                           1 if font_to_use_bold == PDF_FONT_BOLD_NAME else 0, 
                           0, 
                           PDF_FONT_BOLD_NAME) 

                current_app.logger.info(f"Выполнен addMapping для {PDF_FONT_REGULAR_NAME} с жирным {PDF_FONT_BOLD_NAME}")

        else:
            current_app.logger.warning(f"НЕ НАЙДЕН основной файл шрифта: {path_regular_ttf}. Используется Helvetica.")

    except Exception as e:
        current_app.logger.error(f"ОШИБКА при регистрации шрифтов: {e}", exc_info=True)
        font_to_use_regular = 'Helvetica'
        font_to_use_bold = 'Helvetica-Bold'


    report = Report.query.get_or_404(report_id)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30)
    elements = []

    styles = getSampleStyleSheet()

    style_title = ParagraphStyle('TitleStyle', parent=styles['h1'], fontName=font_to_use_regular, alignment=1, fontSize=16, leading=20)
    style_normal = ParagraphStyle('NormalStyle', parent=styles['Normal'], fontName=font_to_use_regular, fontSize=10, leading=12)
    style_bold_table_header = ParagraphStyle('BoldTableHeader', parent=style_normal, fontName=font_to_use_bold, alignment=1, fontSize=10)
    style_table_cell = ParagraphStyle('TableCell', parent=style_normal, fontName=font_to_use_regular, fontSize=9)
    style_table_cell_right = ParagraphStyle('TableCellRight', parent=style_table_cell, alignment=2)


    if report.report_type == 'material_stock' and report.config:
        report_data_config = report.config
        elements.append(Paragraph(report_data_config.get('title', f"Отчет об остатках материалов #{report.id}"), style_title))
        elements.append(Spacer(1, 0.2*72))
        elements.append(Paragraph(f"Сформирован: {report.created_at.strftime('%d.%m.%Y %H:%M')} пользователем {report.user.full_name if report.user else 'Система'}", style_normal))
        elements.append(Paragraph(f"Общая стоимость остатков: {report_data_config.get('total_stock_value', 0):.2f} ₽", style_normal))
        elements.append(Spacer(1, 0.2*72))

        data = []
        header_texts = ['Код', 'Наименование', 'Ед.изм.', 'Остаток', 'Ср.цена', 'Стоимость']
        header_paragraphs = [Paragraph(text, style_bold_table_header) for text in header_texts]
        data.append(header_paragraphs)

        materials_data = report_data_config.get('materials', [])
        for mat in materials_data:
            row = [
                Paragraph(str(mat.get('code', '-')), style_table_cell),
                Paragraph(str(mat.get('name', '-')), style_table_cell),
                Paragraph(str(mat.get('unit', '-')), style_table_cell),
                Paragraph(f"{mat.get('stock_quantity', 0):.3f}", style_table_cell_right),
                Paragraph(f"{mat.get('average_cost_price', 0):.2f}", style_table_cell_right),
                Paragraph(f"{mat.get('total_value', 0):.2f}", style_table_cell_right)
            ]
            data.append(row)
        
        table = Table(data, colWidths=[50, 170, 40, 70, 70, 80])
        table.setStyle(TableStyle([
           ('TEXTCOLOR', (0, 0), (-1, 0), colors.black), 
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black), 
            
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'), 
            ('ALIGN', (0, 1), (2, -1), 'LEFT'),   
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'), 

            ('FONTNAME', (0, 0), (-1, 0), font_to_use_bold),
            ('FONTNAME', (0, 1), (-1, -1), font_to_use_regular),
            
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            
            ('BOTTOMPADDING', (0, 0), (-1, 0), 5), 
            ('TOPPADDING', (0, 0), (-1, 0), 5),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
           
            
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("Неподдерживаемый тип отчета для PDF или данные отчета отсутствуют.", style_normal))

    try:
        doc.build(elements)
    except Exception as e:
        current_app.logger.error(f"Ошибка при сборке PDF документа: {e}", exc_info=True)
        flash("Произошла ошибка при генерации PDF отчета.", "danger")
        return redirect(url_for('routes.report_list'))

    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"report_{report.id}_{report.report_type}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf",
        mimetype="application/pdf"
    )

@bp.route('/reports/<int:report_id>/export/excel')
@login_required
def export_report_excel(report_id):
    if current_user.role not in ['admin', 'accountant', 'analyst']:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.report_list'))
        
    report = Report.query.get_or_404(report_id)
    
    wb = Workbook()
    ws = wb.active
    
    bold_font = Font(bold=True)
    
    if report.report_type == 'material_stock' and report.config: 
        report_data_config = report.config 
        ws.title = f"Остатки материалов {report.id}"
        
        title_cell = ws.cell(row=1, column=1, value=report_data_config.get('title', f"Отчет об остатках материалов #{report.id}"))
        title_cell.font = bold_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6) 

        ws.append(["Сформирован:", report.created_at.strftime('%d.%m.%Y %H:%M')])
        ws.append(["Пользователь:", report.user.full_name if report.user else "Система"])
        ws.append(["Общая стоимость остатков:", report_data_config.get('total_stock_value', 0)])
        ws.cell(row=ws.max_row, column=2).number_format = '#,##0.00 ₽'
        ws.append([]) 

        headers = ["Код", "Наименование", "Ед. изм.", "Остаток на складе", "Ср. цена за ед.", "Общая стоимость"]
        ws.append(headers)
        for cell in ws[ws.max_row]: cell.font = bold_font

        materials_data = report_data_config.get('materials', [])
        for mat in materials_data:
            ws.append([
                mat.get('code', '-'),
                mat.get('name', '-'),
                mat.get('unit', '-'),
                mat.get('stock_quantity', 0),
                mat.get('average_cost_price', 0),
                mat.get('total_value', 0)
            ])
            ws.cell(row=ws.max_row, column=4).number_format = '#,##0.000' 
            ws.cell(row=ws.max_row, column=5).number_format = '#,##0.00 ₽' 
            ws.cell(row=ws.max_row, column=6).number_format = '#,##0.00 ₽' 
            
        ws.append([]) 

        ws.append(["ИТОГО ОБЩАЯ СТОИМОСТЬ:", "", "", "", "", report_data_config.get('total_stock_value', 0)])
        ws.cell(row=ws.max_row, column=1).font = bold_font
        ws.cell(row=ws.max_row, column=6).font = bold_font
        ws.cell(row=ws.max_row, column=6).number_format = '#,##0.00 ₽'
    
    else:
        ws.append(["Неподдерживаемый тип отчета для Excel или данные отчета отсутствуют."])

    for col_idx, column_cells in enumerate(ws.columns, 1): 
        max_length = 0
        column_letter = get_column_letter(col_idx) 

        for cell in column_cells:
  
            if isinstance(cell, MergedCell):
                continue 
            try:
                if cell.value:
 
                    if isinstance(cell.value, (int, float)) and cell.number_format and '₽' in cell.number_format:
     
                        cell_length = len(f"{cell.value:{cell.number_format.replace('₽','').strip()}} ₽")
                    elif isinstance(cell.value, (int, float)):
                        cell_length = len(str(round(cell.value, 3))) + 2 
                    else:
                        cell_length = len(str(cell.value))
                    
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = (max_length + 2) 
        if adjusted_width > 70: 
             adjusted_width = 70
        elif adjusted_width < 10 and headers: 
            adjusted_width = max(10, len(str(ws.cell(row= (5 if report_data_config else 1) , column=col_idx).value)) + 2 if ws.cell(row= (5 if report_data_config else 1), column=col_idx).value else 10)


        ws.column_dimensions[column_letter].width = adjusted_width

    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"report_{report.id}_{report.report_type}_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@bp.route('/materials/categories')
@login_required
def material_categories():
    if current_user.role not in ['admin', 'accountant']:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    categories = MaterialCategory.query.all()
    return render_template('material_categories.html', categories=categories)

@bp.route('/materials/categories/add', methods=['GET', 'POST'])
@login_required
def add_material_category():
    if current_user.role not in ['admin', 'accountant']:
        flash('Доступ запрещен', 'danger')
        current_app.logger.warning(f"User {current_user.username} tried to access add_material_category without permission.")
        return redirect(url_for('routes.dashboard'))
    
    form = MaterialCategoryForm()
    
    if form.validate_on_submit(): 
        current_app.logger.info(f"--- ADD_MAT_CAT (validate_on_submit=True): User {current_user.username} trying to add code '{form.code.data}' ---")
        existing_category = MaterialCategory.query.filter_by(code=form.code.data).first()
        
        if existing_category:
            flash_message = f'Категория с кодом "{form.code.data}" уже существует.'
            flash(flash_message, 'warning')
            current_app.logger.warning(f"--- ADD_MAT_CAT: Duplicate code found. Flash: '{flash_message}' ---")
            return render_template('add_material_category.html', form=form, title="Добавление категории материала")
        
        try:
            category = MaterialCategory(
                name=form.name.data,
                code=form.code.data,
                unit=form.unit.data,
                description=form.description.data,
                stock_quantity=form.stock_quantity.data
            )
            db.session.add(category)
            db.session.commit()
            flash('Категория материала успешно добавлена', 'success') 
            current_app.logger.info(f"--- ADD_MAT_CAT: Category '{form.code.data}' added successfully. Redirecting. ---")
            return redirect(url_for('routes.material_categories'))
        except Exception as e:
            db.session.rollback()
            error_message = f'Ошибка при добавлении категории: {str(e)}'
            flash(error_message, 'danger')
            current_app.logger.error(f"--- ADD_MAT_CAT: Database error on add: {str(e)} ---", exc_info=True)
            return render_template('add_material_category.html', form=form, title="Добавление категории материала")
    else:
        if request.method == 'POST': 
            current_app.logger.error(f"--- ADD_MAT_CAT (POST, validate_on_submit=False): Form errors: {form.errors} for user {current_user.username} ---")
            
    return render_template('add_material_category.html', form=form, title="Добавление категории материала")

@bp.route('/materials/categories/<int:category_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_material_category(category_id):
    if current_user.role not in ['admin', 'accountant']:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    category = MaterialCategory.query.get_or_404(category_id)
    form = MaterialCategoryForm(obj=category)
    
    if form.validate_on_submit():
        try:
            category.name = form.name.data
            category.code = form.code.data
            category.unit = form.unit.data
            category.description = form.description.data
            category.stock_quantity = form.stock_quantity.data
            db.session.commit()
            flash('Категория успешно обновлена', 'success')
            return redirect(url_for('routes.material_categories'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Ошибка при обновлении категории: {str(e)}', 'danger')
    
    return render_template('edit_material_category.html', form=form, category=category, title=f"Редактирование: {category.name}")

@bp.route('/materials/categories/<int:category_id>/delete', methods=['POST'])
@login_required
def delete_material_category(category_id):
    if current_user.role not in ['admin']:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    category = MaterialCategory.query.get_or_404(category_id)

    if category.norms.first() or category.receipts.first() or category.associated_products.first() or category.consumptions: 
        flash(f'Категория "{category.name}" используется и не может быть удалена. Сначала удалите все связанные записи (нормы, оприходования, связи с продукцией, ручные списания).', 'danger')
        return redirect(url_for('routes.material_categories'))
        
    try:
        db.session.delete(category)
        db.session.commit()
        flash('Категория успешно удалена', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при удалении категории: {str(e)}', 'danger')
    
    return redirect(url_for('routes.material_categories'))

@bp.route('/products')
@login_required
def product_list():
    if current_user.role not in ['admin', 'accountant', 'analyst']:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.dashboard'))
    
    products = Product.query.order_by(Product.name).all()
    return render_template('product_list.html', products=products)

@bp.route('/products/add', methods=['GET', 'POST'])
@login_required
def add_product():
    if current_user.role not in ['admin', 'accountant']:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.product_list'))

    form = ProductForm()

    if form.validate_on_submit():
        try:
            new_product = Product(
                name=form.name.data,
                description=form.description.data,
                price=form.price.data,
                length=form.length.data,
                width=form.width.data,
                height=form.height.data,
                dimension_unit=form.dimension_unit.data
            )
            
            db.session.add(new_product)
            db.session.commit()
            flash(f'Продукт "{new_product.name}" успешно добавлен! Теперь определите его спецификацию.', 'success') 
            return redirect(url_for('routes.add_specification_for_product', product_id=new_product.id)) 
        except Exception as e:
            db.session.rollback()
            if "UNIQUE constraint failed: product.name" in str(e).lower():
                 flash('Продукт с таким названием уже существует.', 'danger')
            else:
                flash(f'Ошибка при добавлении продукта: {str(e)}', 'danger')
    
    return render_template('add_edit_product.html', form=form, title="Добавить новый продукт")

@bp.route('/products/<int:product_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_product(product_id):
    if current_user.role not in ['admin', 'accountant']:
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.product_list'))

    product = Product.query.get_or_404(product_id)
    form = ProductForm(obj=product) 

    if form.validate_on_submit():
        try:
            product.name = form.name.data
            product.description = form.description.data
            product.price = form.price.data
            product.length = form.length.data
            product.width = form.width.data
            product.height = form.height.data
            product.dimension_unit = form.dimension_unit.data
            
            db.session.commit()
            flash(f'Продукт "{product.name}" успешно обновлен!', 'success')
            return redirect(url_for('routes.product_list'))
        except Exception as e:
            db.session.rollback()
            if "UNIQUE constraint failed: product.name" in str(e).lower() and \
               not (Product.query.filter(Product.name == form.name.data, Product.id != product_id).first() is None): 
                 flash('Продукт с таким названием уже существует.', 'danger')
            else:
                flash(f'Ошибка при обновлении продукта: {str(e)}', 'danger')

    return render_template('add_edit_product.html', form=form, product=product, title=f"Редактировать продукт: {product.name}")

@bp.route('/products/<int:product_id>/delete', methods=['POST'])
@login_required
def delete_product(product_id):
    if current_user.role not in ['admin']: 
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('routes.product_list'))

    product = Product.query.get_or_404(product_id)
 
    if ProductionLog.query.filter_by(product_id=product.id).first():
        flash(f'Продукт "{product.name}" используется в журнале производства и не может быть удален. Сначала удалите связанные записи о производстве.', 'danger')
        return redirect(url_for('routes.product_list'))
        
    try:
        product.material_categories = []
        db.session.flush()

        db.session.delete(product)
        db.session.commit()
        flash('Продукт успешно удален!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при удалении продукта: {str(e)}', 'danger')
        
    return redirect(url_for('routes.product_list'))

@bp.route('/specifications')
@login_required
def specification_list():
    if current_user.role not in ['admin', 'accountant', 'analyst']:
        flash('Доступ запрещен.', 'danger')
        return redirect(url_for('routes.dashboard'))

    page = request.args.get('page', 1, type=int)
    product_id_filter = request.args.get('product_id_filter', None, type=int) 

    query = Product.query.options(
        db.selectinload(Product.specifications).selectinload(MaterialNorm.material_category)
    )

    if product_id_filter: 
        query = query.filter(Product.id == product_id_filter)
    
    products_with_specs_paginated = query.order_by(Product.name).paginate(page=page, per_page=10 if not product_id_filter else 1) 
    
    if product_id_filter and not products_with_specs_paginated.items:
        filtered_product = Product.query.get(product_id_filter)
        if filtered_product:
             flash(f'Для продукта "{filtered_product.name}" спецификации не найдены.', 'info')
        else:
             flash(f'Продукт с ID {product_id_filter} не найден.', 'warning')

    return render_template('specification_list.html', products_with_specs=products_with_specs_paginated, product_id_filter=product_id_filter)

@bp.route('/specifications/<int:norm_id>/delete', methods=['POST'])
@login_required
def delete_specification(norm_id):
    if current_user.role not in ['admin']:
        flash('Доступ запрещен.', 'danger')
        return redirect(url_for('routes.specification_list'))

    norm = MaterialNorm.query.get_or_404(norm_id)
    product_id_for_redirect = norm.product_id 
    try:
        db.session.delete(norm)
        db.session.commit()
        flash('Норма из спецификации успешно удалена.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при удалении нормы: {str(e)}', 'danger')
    return redirect(url_for('routes.specification_list', product_id_filter=product_id_for_redirect)) 


@bp.route('/production/produce', methods=['GET', 'POST'])
@login_required
def produce_product():
    if current_user.role not in ['admin', 'accountant']: 
        flash('Доступ запрещен.', 'danger')
        return redirect(url_for('routes.dashboard'))

    form = ProduceProductForm()
    available_materials = MaterialCategory.query.all() 

    if form.validate_on_submit():
        product_id = form.product_id.data
        quantity_to_produce = form.quantity_to_produce.data
        production_date = form.production_date.data 
        notes = form.notes.data

        product_to_produce = Product.query.get(product_id)

        if not product_to_produce:
            flash('Выбранный продукт не найден.', 'danger')
            return render_template('produce_product.html', form=form, available_materials=available_materials, title="Производство продукции")

        if not product_to_produce.specifications: 
            flash(f'У продукта "{product_to_produce.name}" нет определенной спецификации. Невозможно произвести.', 'warning')
            return render_template('produce_product.html', form=form, available_materials=available_materials, title="Производство продукции")
        
        try:
            materials_sufficient = True
            required_materials_info = [] 
            total_material_cost_for_batch = 0.0
            material_details_for_log = [] 

            for norm in product_to_produce.specifications:
                material_category = norm.material_category
                required_amount_for_one = norm.norm_value
                total_required_for_batch = required_amount_for_one * quantity_to_produce
                
                current_avg_cost = material_category.average_cost_price
                if current_avg_cost is None or current_avg_cost < 0: 
                    flash(f'ВНИМАНИЕ: У материала "{material_category.name}" не определена или некорректна средняя себестоимость ({current_avg_cost}). Стоимость для этого материала будет рассчитана как 0.', 'warning')
                    current_avg_cost = 0.0

                cost_for_this_material_in_batch = total_required_for_batch * current_avg_cost
                total_material_cost_for_batch += cost_for_this_material_in_batch

                material_details_for_log.append({
                    'material_id': material_category.id,
                    'material_name': material_category.name,
                    'quantity_used': round(total_required_for_batch, 5), 
                    'unit': material_category.unit,
                    'cost_per_unit': round(current_avg_cost, 2),
                    'total_material_cost': round(cost_for_this_material_in_batch, 2)
                })
                
                required_materials_info.append(f"{material_category.name}: требуется {total_required_for_batch:.3f} {material_category.unit} (на складе: {material_category.stock_quantity:.3f} {material_category.unit}, ср.цена: {current_avg_cost:.2f} ₽)")

                if material_category.stock_quantity < total_required_for_batch:
                    materials_sufficient = False
                    flash(f'Недостаточно материала "{material_category.name}"! Требуется: {total_required_for_batch:.3f}, в наличии: {material_category.stock_quantity:.3f}.', 'danger')
            
            if not materials_sufficient:
                return render_template('produce_product.html', form=form, available_materials=available_materials, title="Производство продукции", required_materials_info=required_materials_info)
            
            cost_per_unit_produced = (total_material_cost_for_batch / quantity_to_produce) if quantity_to_produce > 0 else 0.0
            
            new_production_log = ProductionLog(
                product_id=product_to_produce.id,
                quantity_produced=quantity_to_produce,
                production_date=production_date,
                cost_per_unit_calculated=cost_per_unit_produced,
                total_cost_calculated=total_material_cost_for_batch,
                config=material_details_for_log, 
                notes=notes,
                user_id=current_user.id
            )
            db.session.add(new_production_log)
            
            detailed_consumption_info_for_flash = [] 
            for norm in product_to_produce.specifications:
                material_category = norm.material_category 
                total_required_for_batch_for_norm = norm.norm_value * quantity_to_produce
                
                material_category.stock_quantity -= total_required_for_batch_for_norm
                detailed_consumption_info_for_flash.append(f"- {material_category.name}: списано {total_required_for_batch_for_norm:.3f} {material_category.unit} (остаток: {material_category.stock_quantity:.3f} {material_category.unit})")

            db.session.commit()
            
            flash(f'{quantity_to_produce} ед. продукта "{product_to_produce.name}" успешно произведено. Общая материальная себестоимость: {total_material_cost_for_batch:.2f} ₽ (Себестоимость за ед.: {cost_per_unit_produced:.2f} ₽).', 'success')
            
            if detailed_consumption_info_for_flash: 
                flash_message_consumption = "Детализация списания материалов:<br>" + "<br>".join(detailed_consumption_info_for_flash)
                flash(Markup(flash_message_consumption), 'info') 

            return redirect(url_for('routes.production_journal')) 

        except Exception as e:
            db.session.rollback()
            flash(f'Ошибка при производстве продукции: {str(e)}', 'danger')
            current_app.logger.error(f"Production error: {e}", exc_info=True)

    return render_template('produce_product.html', form=form, available_materials=available_materials, title="Производство продукции")


@bp.route('/specifications/add_general', methods=['GET', 'POST'], endpoint='add_specification_general')
@login_required
def add_specification_general_route():
    if current_user.role not in ['admin', 'accountant']:
        flash('Доступ запрещен.', 'danger')
        return redirect(url_for('routes.specification_list'))

    form = MaterialNormForm()

    all_material_categories_query = MaterialCategory.query.order_by(MaterialCategory.name).all()
 
    all_material_categories = all_material_categories_query if all_material_categories_query is not None else []

    all_material_categories_json_data = [cat.to_dict() for cat in all_material_categories]

    material_units_dict = {str(cat.id): (cat.unit if cat.unit and cat.unit.strip() else 'ед.')
                           for cat in all_material_categories}

    missing_products = not form.product_id.choices or \
                       (len(form.product_id.choices) == 1 and form.product_id.choices[0][0] == 0)
    missing_categories = not form.material_category_id.choices or \
                         (len(form.material_category_id.choices) == 1 and form.material_category_id.choices[0][0] == 0)

    template_context = {
        "form": form,
        "title": "Добавить норму в спецификацию",
        "missing_products": missing_products,
        "missing_categories": missing_categories,
        "all_material_categories": all_material_categories,
        "material_units_map_for_js": material_units_dict,
        "all_material_categories_json_data": all_material_categories_json_data, 
        "product_id_for_form_action": None,
        "specification": None
    }

    if missing_products or missing_categories:
        if request.method == 'POST':
             flash('Необходимо сначала добавить продукты и/или категории материалов.', 'warning')
        return render_template('add_edit_specification.html', **template_context)


    if form.validate_on_submit():
        existing_norm = MaterialNorm.query.filter_by(
            product_id=form.product_id.data,
            material_category_id=form.material_category_id.data
        ).first()
        if existing_norm:
            flash('Норма для этого продукта и материала уже существует. Отредактируйте существующую.', 'warning')
        else:
            try:
                norm = MaterialNorm(
                    product_id=form.product_id.data,
                    material_category_id=form.material_category_id.data,
                    norm_value=form.norm_value.data
                )
                db.session.add(norm)
                db.session.commit()
                flash('Норма успешно добавлена в спецификацию!', 'success')
                return redirect(url_for('routes.specification_list', product_id_filter=form.product_id.data))
            except Exception as e:
                db.session.rollback()
                flash(f'Ошибка при добавлении нормы: {str(e)}', 'danger')

        return render_template('add_edit_specification.html', **template_context)

    return render_template('add_edit_specification.html', **template_context)

@bp.route('/product/<int:product_id>/specifications/add', methods=['GET', 'POST'], endpoint='add_specification_for_product')
@login_required
def add_specification_for_product_route(product_id):
    if current_user.role not in ['admin', 'accountant']:
        flash('Доступ запрещен.', 'danger')
        return redirect(url_for('routes.specification_list'))
    
    product = Product.query.get_or_404(product_id)
    form = MaterialNormForm()
    
    all_material_categories = MaterialCategory.query.order_by(MaterialCategory.name).all()
    all_material_categories_json_data = [cat.to_dict() for cat in all_material_categories]
    material_units_dict = {str(cat.id): (cat.unit if cat.unit and cat.unit.strip() else 'ед.') 
                           for cat in all_material_categories}
    
                           
    form.product_id.data = product_id 
    if form.product_id.render_kw is None: form.product_id.render_kw = {}
    form.product_id.render_kw['disabled'] = True
    
    current_product_choice = (product.id, product.name)
    if form.product_id.choices and form.product_id.choices[0][0] == 0: 
        form.product_id.choices.pop(0)
    if current_product_choice not in form.product_id.choices:
        form.product_id.choices.insert(0, current_product_choice)
    form.product_id.choices = sorted([c for c in form.product_id.choices if c[0] != 0] + ([current_product_choice] if current_product_choice not in form.product_id.choices else []), key=lambda x:x[1])


    missing_categories = not form.material_category_id.choices or \
                         (len(form.material_category_id.choices) == 1 and form.material_category_id.choices[0][0] == 0)
    if missing_categories:
        flash('Необходимо сначала добавить категории материалов.', 'warning')
        return render_template('add_edit_specification.html', form=form, title=f"Добавить норму для \"{product.name}\"", 
                               missing_categories=missing_categories, missing_products=False, 
                               product_id_for_form_action=product_id,
                               all_material_categories=all_material_categories,
                               material_units_map_for_js=material_units_dict) 

    if form.validate_on_submit():
        existing_norm = MaterialNorm.query.filter_by(
            product_id=product_id, 
            material_category_id=form.material_category_id.data
        ).first()
        if existing_norm:
            flash('Норма для этого продукта и материала уже существует. Отредактируйте существующую.', 'warning')
        else:
            try:
                norm = MaterialNorm(
                    product_id=product_id, 
                    material_category_id=form.material_category_id.data,
                    norm_value=form.norm_value.data
                )
                db.session.add(norm)
                db.session.commit()
                flash('Норма успешно добавлена в спецификацию!', 'success')
                return redirect(url_for('routes.specification_list', product_id_filter=product_id))
            except Exception as e:
                db.session.rollback()
                flash(f'Ошибка при добавлении нормы: {str(e)}', 'danger')
    
    return render_template('add_edit_specification.html', form=form, title=f"Добавить норму для \"{product.name}\"", 
                           missing_categories=missing_categories, missing_products=False, 
                           product_id_for_form_action=product_id,
                           all_material_categories=all_material_categories,
                           material_units_map_for_js=material_units_dict) 


@bp.route('/specifications/<int:norm_id>/edit', methods=['GET', 'POST'], endpoint='edit_specification')
@login_required
def edit_specification_route(norm_id):
    if current_user.role not in ['admin', 'accountant']:
        flash('Доступ запрещен.', 'danger')
        return redirect(url_for('routes.specification_list'))

    norm = MaterialNorm.query.get_or_404(norm_id)
    form = MaterialNormForm(obj=norm) 
    
    all_material_categories = MaterialCategory.query.order_by(MaterialCategory.name).all()
    all_material_categories_json_data = [cat.to_dict() for cat in all_material_categories]
    material_units_dict = {str(cat.id): (cat.unit if cat.unit and cat.unit.strip() else 'ед.') 
                           for cat in all_material_categories}

    if form.product_id.render_kw is None: form.product_id.render_kw = {}
    form.product_id.render_kw['disabled'] = True
    if form.material_category_id.render_kw is None: form.material_category_id.render_kw = {}
    form.material_category_id.render_kw['disabled'] = True

    if form.validate_on_submit():
        try:
            norm.norm_value = form.norm_value.data
            db.session.commit()
            flash('Норма в спецификации успешно обновлена!', 'success')
            return redirect(url_for('routes.specification_list', product_id_filter=norm.product_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Ошибка при обновлении нормы: {str(e)}', 'danger')
            
    return render_template('add_edit_specification.html', form=form, title=f"Редактировать норму для \"{norm.product.name}\"", 
                           specification=norm, 
                           missing_products=False, missing_categories=False,
                           all_material_categories=all_material_categories,
                           material_units_map_for_js=material_units_dict) 

@bp.route('/materials/receipts/add', methods=['GET', 'POST'], endpoint='add_material_receipt') 
@login_required
def add_material_receipt(): 
    if current_user.role not in ['admin', 'accountant']:
        flash('Доступ запрещен.', 'danger')
        return redirect(url_for('routes.material_receipt_list'))     

    form = MaterialReceiptForm() 

    all_material_categories = MaterialCategory.query.order_by(MaterialCategory.name).all()
    material_units_dict = {
        str(cat.id): (cat.unit if cat.unit and cat.unit.strip() else 'ед.')
        for cat in all_material_categories
    }
    
    if form.validate_on_submit():
        if form.material_category_id.data == 0 or form.material_category_id.data is None: 
            flash('Пожалуйста, выберите категорию материала.', 'warning')
        else:
            try:
                material_category = MaterialCategory.query.get(form.material_category_id.data)
                if not material_category:
                    flash('Выбранная категория материала не найдена.', 'danger')
                else:
                    quantity_received = form.quantity_received.data
                    price_per_unit = form.price_per_unit.data 

                    if price_per_unit is not None and price_per_unit >= 0 and quantity_received > 0:
                        current_stock_before_receipt = material_category.stock_quantity
                        current_avg_cost_before_receipt = material_category.average_cost_price if material_category.average_cost_price is not None else 0.0
                        
                        total_value_before_receipt = current_stock_before_receipt * current_avg_cost_before_receipt
                        value_of_this_receipt = quantity_received * price_per_unit
                        
                        new_total_stock = current_stock_before_receipt + quantity_received
                        
                        if new_total_stock > 0: 
                            material_category.average_cost_price = (total_value_before_receipt + value_of_this_receipt) / new_total_stock
                        elif quantity_received > 0 : 
                            material_category.average_cost_price = price_per_unit
                    
                    receipt = MaterialReceipt(
                        material_category_id=material_category.id,
                        quantity_received=quantity_received, 
                        receipt_date=form.receipt_date.data,        
                        document_ref=form.document_ref.data,
                        supplier=form.supplier.data,
                        price_per_unit=price_per_unit, 
                        notes=form.notes.data,
                        user_id=current_user.id
                    )
                    
                    material_category.stock_quantity += quantity_received 
                    
                    db.session.add(receipt)
                    db.session.commit()
                    
                    flash_message = f'Материал "{material_category.name}" успешно оприходован.'
                    if material_category.average_cost_price is not None:
                        flash_message += f' Новая ср. себестоимость: {material_category.average_cost_price:.2f} ₽.'
                    flash(flash_message, 'success')
                    return redirect(url_for('routes.material_receipt_list')) 
            except Exception as e:
                db.session.rollback()
                flash(f'Ошибка при оприходовании материала: {str(e)}', 'danger')
                current_app.logger.error(f"Material receipt error: {e}", exc_info=True)
    
    return render_template(
        'add_edit_material_receipt.html', 
        form=form, 
        title="Оприходование материала",
        all_material_categories=all_material_categories, 
        material_units_data=material_units_dict  
    )

@bp.route('/materials/receipts', methods=['GET']) 
@login_required
def material_receipt_list(): 
    if current_user.role not in ['admin', 'accountant', 'analyst']: 
        flash('Доступ к истории оприходований запрещен.', 'danger')
        return redirect(url_for('routes.dashboard'))

    page = request.args.get('page', 1, type=int)
    receipts_pagination = MaterialReceipt.query.options(
        db.joinedload(MaterialReceipt.material_category),
        db.joinedload(MaterialReceipt.user)
    ).order_by(
        MaterialReceipt.receipt_date.desc(), MaterialReceipt.created_at.desc()
    ).paginate(page=page, per_page=15) 

    return render_template(
        'material_receipt_list.html', 
        title="История оприходований материалов",
        receipts=receipts_pagination.items,
        pagination=receipts_pagination
    )